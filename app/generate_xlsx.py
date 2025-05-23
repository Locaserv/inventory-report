from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import pdfplumber

from functools import reduce
import logging
import re
import os

logging.getLogger("pdfminer").setLevel(logging.ERROR)

class PDFToExcelConverter:
    def __init__(self, xlsx_folder="inventory_report"):
        self._xlsx_path = xlsx_folder
        self._pdf_paths = []
        self._data = []

    def set_xlsx_path(self, path):
        self._xlsx_path = path
    
    def get_xlsx_path(self):
        return self._xlsx_path

    def add_pdf_path(self, path):
        self._pdf_paths.append(path)
    
    def extract_local_from_pdf(self, pagina):
        """
        Extrai o local diretamente do texto bruto da página.
        """
        texto = pagina.extract_text()
        match = re.search(r"LLooccaall::\s*(\d+\s*-\s*.+)", texto)
        if match:
            return match.group(1).strip().split(" ")[-1].upper()
        return None

    def extract_data(self):
        self._data = []  # reset

        for pdf in self._pdf_paths:
            with pdfplumber.open(pdf) as pdf_file:
                local = self.extract_local_from_pdf(pdf_file.pages[0])
                for page in pdf_file.pages:
                    lines = page.extract_text().split("\n")
                    for line in lines:
                        line = line.strip()
                        # só processa linhas que comecem com 8 dígitos
                        if re.match(r"^\d{8}", line):
                            partes = line.rsplit(maxsplit=3)
                            if len(partes) < 4:
                                continue
                            try:
                                total = float(partes[-1].replace('.', '').replace(',', '.'))
                                preco_medio = float(partes[-2].replace('.', '').replace(',', '.'))
                                saldo = float(partes[-3].replace('.', '').replace(',', '.'))
                                codigo = line[:8]
                                descricao = line[8: -len(partes[-3]) - len(partes[-2]) - len(partes[-1]) - 3].strip()

                                self._data.append({
                                    "Código": codigo,
                                    "Descrição": descricao,
                                    "Local": local,
                                    "Quantidade": saldo,
                                    "Preço Médio": preco_medio,
                                    "Total": total
                                })
                            except Exception:
                                print(f"⚠️ Erro ao processar linha: {line}")
                                continue

        return pd.DataFrame(self._data)

    
    def generate_dataframe(self):
        df_raw = self.extract_data()

        if df_raw.empty:
            print("⚠️ Nenhum dado foi extraído dos PDFs.")
            return None

        # Mapeia a descrição por código
        descricao_por_codigo = df_raw.drop_duplicates(subset="Código")[["Código", "Descrição"]].set_index("Código").to_dict()["Descrição"]

        # Agrupamento
        df_grouped = df_raw.groupby(["Código", "Local"], as_index=False).agg({
            "Quantidade": "sum",
            "Preço Médio": "mean"
        })

        filiais = df_grouped["Local"].unique()
        dfs = []

        for filial in filiais:
            df_filial = df_grouped[df_grouped["Local"] == filial][["Código", "Quantidade"]]
            df_filial = df_filial.rename(columns={
                "Quantidade": f"QUANT {filial}"
            })
            dfs.append(df_filial)

        df_merged = reduce(
            lambda left, right: pd.merge(left, right, on="Código", how="outer"),
            dfs
        )

        # Restaura a descrição
        df_merged["DESCRIÇÃO"] = df_merged["Código"].map(descricao_por_codigo)

        # Preenche valores nulos
        df_merged.fillna(0, inplace=True)

        # Calcula preço máximo entre todas as filiais
        max_price_df = df_grouped.groupby("Código")["Preço Médio"].max().reset_index()
        max_price_df = max_price_df.rename(columns={"Preço Médio": "PREÇO"})

        df_final = pd.merge(df_merged, max_price_df, on="Código", how="left")

        # Soma total de quantidades
        quant_cols = [col for col in df_final.columns if col.startswith("QUANT")]
        df_final["QTD TOTAL"] = df_final[quant_cols].sum(axis=1)

        # Calcula o PREÇO TOTAL
        df_final["PREÇO TOTAL"] = df_final["PREÇO"] * df_final["QTD TOTAL"]

        # Reorganiza as colunas
        colunas_finais = ["Código", "DESCRIÇÃO"] + quant_cols + ["PREÇO", "PREÇO TOTAL"]
        df_final = df_final[colunas_finais]

        return df_final
    
    def save_report_to_excel(self, df):
        if df is None:
            return None

        os.makedirs(self._xlsx_path, exist_ok=True)
        filename = f"relatorio_estoque_{pd.Timestamp.now().strftime('%d%m%Y_%H%M%S')}.xlsx"
        output_path = os.path.join(self._xlsx_path, filename)

        df.to_excel(output_path, index=False, engine="openpyxl", startrow=2)

        wb = load_workbook(output_path)
        ws = wb.active

        # Header
        today = pd.Timestamp.now().strftime("%d/%m/%Y")
        ws.merge_cells("A1:I1")
        ws["A1"] = "LOCASERV - LOCAÇÃO E SERVIÇOS LTDA"
        ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
        ws["A1"].fill = PatternFill("solid", fgColor="4F81BD")
        ws["A1"].alignment = Alignment(horizontal="center")

        # Sub-Header
        ws.merge_cells("A2:I2")
        ws["A2"] = f"RELATÓRIO GERAL DE ESTOQUE    DATA: {today}"
        ws["A2"].font = Font(size=12, italic=True, bold=True, color="FFFFFF")
        ws["A2"].fill = PatternFill("solid", fgColor="4F81BD")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Header table
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")

        for col_num, cell in enumerate(ws[3], 1):
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for col_cells in ws.iter_cols(min_row=3, max_row=ws.max_row):
            header_value = str(col_cells[0].value) if col_cells[0].value else ""
            max_length = len(header_value)


            for cell in col_cells[1:]:
                if isinstance(cell.value, (float, int)):
                    cell.number_format = '#,##0.00'
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))

            col_letter = get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = max_length + 4

        wb.save(output_path)

        return output_path
    
    def generate_report(self):
        df = self.generate_dataframe()
        return self.save_report_to_excel(df)


# How to use the class
# from pdf_to_excel import PDFToExcelConverter

# converter = PDFToExcelConverter(xlsx_folder="./relatorios")
# converter.add_pdf_path("./entrada/PETROLINA.pdf")
# converter.add_pdf_path("./entrada/GARANHUNS.pdf")
# converter.add_pdf_path("./entrada/CRUZ_DE_SALINAS.pdf")
# converter.add_pdf_path("./entrada/IPOJUCA.pdf")
# converter.add_pdf_path("./entrada/RAJADA.pdf")

# converter.set_xlsx_path("relatorios")
# converter.generate_report()